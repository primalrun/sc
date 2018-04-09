import os
import sys
import shutil
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook

 
#Requires c:\qv_dev\qvd_list.txt to be updated with qvd files for *dashboard.qvw
local_dir = 'C:/QlikView/'
qvd_dir = local_dir + 'QVD'
s_file = local_dir + r'qvd_list.txt'
prod_dir = '//QlikView/C$/QlikView/'
qvd_prod_dir = prod_dir + 'QVD'
source_data_prod_dir = prod_dir + 'SourceData'

#user select qvw file
root = tk.Tk()
root.withdraw()
qvw_file = filedialog.askopenfilename(title='Select qvw Dashboard File on Production Directory',
                                      initialdir=prod_dir + '/' + 'SourceDocuments')

if len(qvw_file) == 0:
    print('No qvw file selected, try again')
    sys.exit()
qvw_split = qvw_file.split('/')
cat_folder = qvw_split[-2]
qvw_file = qvw_split[-1]

#qvd directory
if not os.path.exists(qvd_dir):
    os.mkdir(qvd_dir)

#sourcedocuments directory
if not os.path.exists(local_dir + '/' + 'SourceDocuments'):    
    os.mkdir(local_dir + '/' + 'SourceDocuments')

#sourcedata directory
if not os.path.exists(local_dir + '/' + 'SourceData'):    
    os.mkdir(local_dir + '/' + 'SourceData')

#sourcedata files    
src_file = prod_dir + 'SourceData' + '/' + 'Qlik Content and Security.xlsx'
dest_file = local_dir + '/' + 'SourceData' + '/' + 'Qlik Content and Security.xlsx'    
shutil.copy2(src=src_file, dst=dest_file, follow_symlinks=False)
print('Copying {} to {}'.format(src_file, dest_file))

#source files for user selected qvw file from Qlik Content and Security.xlsx
source_file = local_dir + '/' + 'SourceData' + '/' + 'Qlik Content and Security.xlsx'
wb_source = load_workbook(source_file)
ws_source = wb_source['Qlik Metadata']

for i in range(1, 5):
    if ws_source.cell(row=2, column=i).value == 'Application (QVW)':
        qvw_col = i
        r1 = 3
        break
if qvw_col is None:
    print('Unable to find column "Application (QVW)" in file Qlik Content and Security.xlsx, Process Cancelled')
    sys.exit()

for i in range(1, 5):
    if ws_source.cell(row=2, column=i).value == 'First Level Source Data (QVD)':
        source_col = i
        break
if source_col is None:
    print('Unable to find column "First Level Source Data (QVD)" in file Qlik Content and Security.xlsx, Process Cancelled')
    sys.exit()

r_last = len(ws_source['A'])
qvd_dict = {}
for r in range(r1, r_last+1):
    if ws_source.cell(row=r, column=qvw_col).value == qvw_file:
        qvd_dict[ws_source.cell(row=r, column=source_col).value] = None
        
wb_source.close()
  
#populate list with directory, folder name, qvd file
qvd_file_list = []
source_data_file_list = []

#qvd files
dir_list = os.listdir(qvd_prod_dir)
for v in qvd_dict:
    for f in dir_list:        
        if os.path.isfile(qvd_prod_dir + '/' + f + '/' + v):
            qvd_file_list.append((qvd_prod_dir, f, v))
            break

#sourcedata files
dir_list = os.listdir(source_data_prod_dir)
for v in qvd_dict:            
    if os.path.isfile(source_data_prod_dir + '/' + v):
        source_data_file_list.append((prod_dir + '/', 'SourceData', v))
        break
    
#create folders on local pc
for f in qvd_file_list:    
    if not os.path.exists(qvd_dir + '/' + f[1]):
        os.mkdir(qvd_dir + '/' + f[1])

if not os.path.exists(local_dir + '/' + 'SourceData'):
    os.mkdir(local_dir + '/' + 'SourceData')

#copy qvd files from production to local pc
for f in qvd_file_list:
    src_file = f[0] + '/' + f[1] + '/' +f[2]
    dest_file = qvd_dir + '/' + f[1] + '/' +f[2]
    shutil.copy2(src=src_file, dst=dest_file, follow_symlinks=False)
    print('Copying {} to {}'.format(src_file, dest_file))

#copy source data files from production to local pc
for f in source_data_file_list:
    src_file = f[0] + '/' + f[1] + '/' +f[2]
    dest_file = local_dir + '/' + f[1] + '/' +f[2]
    shutil.copy2(src=src_file, dst=dest_file, follow_symlinks=False)
    print('Copying {} to {}'.format(src_file, dest_file))

#copy other general purpose source files
#qvd_history
if not os.path.exists(qvd_dir + '/' + 'Administrative'):
    os.mkdir(qvd_dir + '/' + 'Administrative')
src_file = qvd_prod_dir + '/' + 'Administrative' + '/' + 'qvd_history.qvd'
dest_file = qvd_dir + '/' + 'Administrative' + '/' + 'qvd_history.qvd'    
shutil.copy2(src=src_file, dst=dest_file, follow_symlinks=False)
print('Copying {} to {}'.format(src_file, dest_file))

#sourcedocuments files
if not os.path.exists(local_dir + '/' + 'SourceDocuments'):
    os.mkdir(local_dir + '/' + 'SourceDocuments')
if not os.path.exists(local_dir + '/' + 'SourceDocuments' + '/' + '_Includes'):
    os.mkdir(local_dir + '/' + 'SourceDocuments' + '/' + '_Includes')

src_file = prod_dir + 'SourceDocuments' + '/' + '_Includes' + '/' + 'qvdmaker.txt'
dest_file = local_dir + '/' + 'SourceDocuments' + '/' + '_Includes' + '/' + 'qvdmaker.txt'    
shutil.copy2(src=src_file, dst=dest_file, follow_symlinks=False)
print('Copying {} to {}'.format(src_file, dest_file))

src_file = prod_dir + 'SourceDocuments' + '/' + '_Includes' + '/' + 'qvdreloadtimes.txt'
dest_file = local_dir + '/' + 'SourceDocuments' + '/' + '_Includes' + '/' + 'qvdreloadtimes.txt'    
shutil.copy2(src=src_file, dst=dest_file, follow_symlinks=False)
print('Copying {} to {}'.format(src_file, dest_file))

src_file = prod_dir + 'SourceDocuments' + '/' + '_Includes' + '/' + 'qvdreloadtimessense.txt'
dest_file = local_dir + '/' + 'SourceDocuments' + '/' + '_Includes' + '/' + 'qvdreloadtimessense.txt'    
shutil.copy2(src=src_file, dst=dest_file, follow_symlinks=False)
print('Copying {} to {}'.format(src_file, dest_file))

src_file = prod_dir + 'SourceDocuments' + '/' + '_Includes' + '/' + 'sc_color_palettes.qvs'
dest_file = local_dir + '/' + 'SourceDocuments' + '/' + '_Includes' + '/' + 'sc_color_palettes.qvs'    
shutil.copy2(src=src_file, dst=dest_file, follow_symlinks=False)
print('Copying {} to {}'.format(src_file, dest_file))

src_file = prod_dir + 'SourceDocuments' + '/' + '_Includes' + '/' + 'sc_master_calendar.qvs'
dest_file = local_dir + '/' + 'SourceDocuments' + '/' + '_Includes' + '/' + 'sc_master_calendar.qvs'    
shutil.copy2(src=src_file, dst=dest_file, follow_symlinks=False)
print('Copying {} to {}'.format(src_file, dest_file))

#qvw file selected by user
if not os.path.exists(local_dir + '/' + 'SourceDocuments' + '/' + cat_folder):
    os.mkdir(local_dir + '/' + 'SourceDocuments' + '/' + cat_folder)

src_file = prod_dir + 'SourceDocuments' + '/' + cat_folder + '/' + qvw_file
dest_file = local_dir +  'SourceDocuments' + '/' + cat_folder + '/' + qvw_file    
shutil.copy2(src=src_file, dst=dest_file, follow_symlinks=False)
print('Copying {} to {}'.format(src_file, dest_file))

    
print('Success, Press Enter to close command prompt')
input()
