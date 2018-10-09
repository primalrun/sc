import os
import sys
import shutil
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
import time

#find file function
def find_file(fname, fpath):    
    for root, dirs, files in os.walk(fpath):
        if fname in files:
            return os.path.join(root, fname)
        
#command line resize
os.system('mode con cols=200 lines=50')
 
local_dir = 'C:/QlikView/'
qvd_dir = local_dir + 'QVD'
prod_dir = '//QlikView/C$/QlikView/'
qvd_prod_dir = prod_dir + 'QVD'
source_data_prod_dir = prod_dir + 'SourceData'

#user select qvw file
root = tk.Tk()
root.withdraw()
qvw_file = filedialog.askopenfilename(title='Select qvw Dashboard File on Production Directory',
                                      initialdir=prod_dir + '/' + 'SourceDocuments')

if len(qvw_file) == 0:
    print('No qvw file selected, try again')
    sys.exit()
qvw_split = qvw_file.split('/')
cat_folder = qvw_split[-2]
qvw_file = qvw_split[-1]
time_stamp=time.strftime('%Y%m%d-%H%M%S')
qvw_file_new = (qvw_file.split('.')[0] + ' ' + 
                time_stamp + '.' + 
                qvw_file.split('.')[1]) 

#qvd directory
if not os.path.exists(qvd_dir):
    os.mkdir(qvd_dir)

#sourcedocuments directory
if not os.path.exists(local_dir + '/' + 'SourceDocuments'):    
    os.mkdir(local_dir + '/' + 'SourceDocuments')

#sourcedata directory
if not os.path.exists(local_dir + '/' + 'SourceData'):    
    os.mkdir(local_dir + '/' + 'SourceData')

#sourcedata files    
src_file = prod_dir + 'SourceData' + '/' + 'Qlik Content and Security.xlsx'
dest_file = local_dir + '/' + 'SourceData' + '/' + 'Qlik Content and Security.xlsx'    
print('Copying {} to {}'.format(src_file, dest_file))
shutil.copy2(src=src_file, dst=dest_file, follow_symlinks=False)

src_file = prod_dir + 'SourceData' + '/' + 'QVD Error Thresholds.xlsx'
dest_file = local_dir + '/' + 'SourceData' + '/' + 'QVD Error Thresholds.xlsx'    
print('Copying {} to {}'.format(src_file, dest_file))
shutil.copy2(src=src_file, dst=dest_file, follow_symlinks=False)


#source files for user selected qvw file from Qlik Content and Security.xlsx
source_file = local_dir + '/' + 'SourceData' + '/' + 'Qlik Content and Security.xlsx'
wb_source = load_workbook(source_file)
ws_source = wb_source['Qlik Metadata']

for i in range(1, 5):
    if ws_source.cell(row=2, column=i).value == 'Application (QVW)':
        qvw_col = i
        r1 = 3
        break
if qvw_col is None:
    print('Unable to find column "Application (QVW)" in file Qlik Content and Security.xlsx, Process Cancelled')
    sys.exit()

for i in range(1, 5):
    if ws_source.cell(row=2, column=i).value == 'First Level Source Data (QVD)':
        source_col = i
        break
if source_col is None:
    print('Unable to find column "First Level Source Data (QVD)" in file Qlik Content and Security.xlsx, Process Cancelled')
    sys.exit()

r_last = len(ws_source['A'])
qvd_dict = {}
for r in range(r1, r_last+1):
    if ws_source.cell(row=r, column=qvw_col).value == qvw_file:
        qvd_dict[ws_source.cell(row=r, column=source_col).value] = None
      
wb_source.close()

#ColorPallette(50).xlsx
src_file = prod_dir + 'SourceData' + '/' + 'ColorPalette(50).xlsx'
dest_file = local_dir + '/' + 'SourceData' + '/' + 'ColorPalette(50).xlsx'    
print('Copying {} to {}'.format(src_file, dest_file))
shutil.copy2(src=src_file, dst=dest_file, follow_symlinks=False)
  
#populate list with directory, folder name, qvd file
qvd_file_list = []
source_data_file_list = []

#copy qvd files
for v in qvd_dict:
    #get production file path
    qvd_file_production = find_file(v, qvd_prod_dir)
    if qvd_file_production is not None:
        #get directory portion only
        qvd_file_path_only = os.path.dirname(qvd_file_production)
        #get file name only
        qvd_filename_only = os.path.basename(qvd_file_production)
        #make directory on local machine if needed        
        local_qvd_directory = os.path.join(qvd_dir, 
                                           qvd_file_path_only.split('QVD\\')[1]
                                           )
        if not os.path.exists(local_qvd_directory):
            os.makedirs(local_qvd_directory)
        #copy qvd files from production to local
        src_file = qvd_file_production
        dest_file = os.path.join(local_qvd_directory, qvd_filename_only)
        print('Copying {} to {}'.format(src_file, dest_file))
        shutil.copy2(src=src_file, dst=dest_file, follow_symlinks=False)

#sourcedata files
dir_list = os.listdir(source_data_prod_dir)
for v in qvd_dict:            
    if os.path.isfile(source_data_prod_dir + '/' + v):
        source_data_file_list.append((prod_dir + '/', 'SourceData', v))
        break
    
#create folders on local pc        
if not os.path.exists(local_dir + '/' + 'SourceData'):
    os.mkdir(local_dir + '/' + 'SourceData')

#copy source data files from production to local pc
for f in source_data_file_list:
    src_file = f[0] + '/' + f[1] + '/' +f[2]
    dest_file = local_dir + '/' + f[1] + '/' +f[2]
    print('Copying {} to {}'.format(src_file, dest_file))
    shutil.copy2(src=src_file, dst=dest_file, follow_symlinks=False)

#copy other general purpose source files
#qvd_history
if not os.path.exists(qvd_dir + '/' + 'Administrative'):
    os.mkdir(qvd_dir + '/' + 'Administrative')
src_file = qvd_prod_dir + '/' + 'Administrative' + '/' + 'qvd_history.qvd'
dest_file = qvd_dir + '/' + 'Administrative' + '/' + 'qvd_history.qvd'    
print('Copying {} to {}'.format(src_file, dest_file))
shutil.copy2(src=src_file, dst=dest_file, follow_symlinks=False)


#sourcedocuments files
if not os.path.exists(local_dir + '/' + 'SourceDocuments'):
    os.mkdir(local_dir + '/' + 'SourceDocuments')
if not os.path.exists(local_dir + '/' + 'SourceDocuments' + '/' + '_Includes'):
    os.mkdir(local_dir + '/' + 'SourceDocuments' + '/' + '_Includes')

src_file = prod_dir + 'SourceDocuments' + '/' + '_Includes' + '/' + 'qvdmaker.txt'
dest_file = local_dir + '/' + 'SourceDocuments' + '/' + '_Includes' + '/' + 'qvdmaker.txt'    
print('Copying {} to {}'.format(src_file, dest_file))
shutil.copy2(src=src_file, dst=dest_file, follow_symlinks=False)


src_file = prod_dir + 'SourceDocuments' + '/' + '_Includes' + '/' + 'qvdreloadtimes.txt'
dest_file = local_dir + '/' + 'SourceDocuments' + '/' + '_Includes' + '/' + 'qvdreloadtimes.txt'    
print('Copying {} to {}'.format(src_file, dest_file))
shutil.copy2(src=src_file, dst=dest_file, follow_symlinks=False)


src_file = prod_dir + 'SourceDocuments' + '/' + '_Includes' + '/' + 'qvdreloadtimessense.txt'
dest_file = local_dir + '/' + 'SourceDocuments' + '/' + '_Includes' + '/' + 'qvdreloadtimessense.txt'    
print('Copying {} to {}'.format(src_file, dest_file))
shutil.copy2(src=src_file, dst=dest_file, follow_symlinks=False)


src_file = prod_dir + 'SourceDocuments' + '/' + '_Includes' + '/' + 'sc_color_palettes.qvs'
dest_file = local_dir + '/' + 'SourceDocuments' + '/' + '_Includes' + '/' + 'sc_color_palettes.qvs'    
print('Copying {} to {}'.format(src_file, dest_file))
shutil.copy2(src=src_file, dst=dest_file, follow_symlinks=False)


src_file = prod_dir + 'SourceDocuments' + '/' + '_Includes' + '/' + 'sc_master_calendar.qvs'
dest_file = local_dir + '/' + 'SourceDocuments' + '/' + '_Includes' + '/' + 'sc_master_calendar.qvs'    
print('Copying {} to {}'.format(src_file, dest_file))
shutil.copy2(src=src_file, dst=dest_file, follow_symlinks=False)


#qvw file selected by user
if not os.path.exists(local_dir + '/' + 'SourceDocuments' + '/' + cat_folder):
    os.mkdir(local_dir + '/' + 'SourceDocuments' + '/' + cat_folder)

src_file = prod_dir + 'SourceDocuments' + '/' + cat_folder + '/' + qvw_file
dest_file = local_dir +  'SourceDocuments' + '/' + cat_folder + '/' + qvw_file_new    
print('Copying {} to {}'.format(src_file, dest_file))
shutil.copy2(src=src_file, dst=dest_file, follow_symlinks=False)
    
print('Success, Press Enter to close command prompt')
input()
